from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import db


def _latest_snapshot_df() -> pd.DataFrame:
    rows = db.dump_history()
    # dump_history() возвращает: (product_id, name, price, checked_at)
    df = pd.DataFrame(rows, columns=["product_id", "name", "price", "checked_at"])
    if df.empty:
        return df

    # сортируем и берём последний замер по каждому товару
    df_sorted = df.sort_values(["product_id", "checked_at"], ascending=[True, True])
    latest_idx = df_sorted.groupby("product_id")["checked_at"].idxmax()
    latest = df_sorted.loc[latest_idx].copy()

    # предыдущая цена (если есть)
    second_idx = (
        df_sorted.groupby("product_id")
        .apply(lambda g: g.iloc[-2].name if len(g) >= 2 else None)
        .dropna()
        .astype(int)
    )
    prev = df_sorted.loc[second_idx] if len(second_idx) else pd.DataFrame(columns=df.columns)

    latest = latest.merge(
        prev[["product_id", "price"]].rename(columns={"price": "prev_price"}),
        on="product_id",
        how="left",
    )
    latest["change"] = latest["price"] - latest["prev_price"]
    latest["change_pct"] = (latest["change"] / latest["prev_price"]) * 100
    return latest


def build_excel_report(out_path: Path) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    latest = _latest_snapshot_df()

    # === Формируем листы ===
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        # Лист «Текущие цены»
        if latest.empty:
            latest_out = pd.DataFrame(
                columns=["Товар", "Цена (₽)", "Предыдущая цена (₽)", "Изменение (₽)", "Изменение (%)", "Когда проверено"]
            )
            latest_out.to_excel(xw, sheet_name="Текущие цены", index=False)
        else:
            latest_out = latest[["name", "price", "prev_price", "change", "change_pct", "checked_at"]].copy()
            latest_out.rename(
                columns={
                    "name": "Товар",
                    "price": "Цена (₽)",
                    "prev_price": "Предыдущая цена (₽)",
                    "change": "Изменение (₽)",
                    "change_pct": "Изменение (%)",
                    "checked_at": "Когда проверено",
                },
                inplace=True,
            )
            latest_out.to_excel(xw, sheet_name="Текущие цены", index=False)

        # Лист «История»
        hist_rows = db.dump_history()
        hist_df = pd.DataFrame(hist_rows, columns=["product_id", "name", "price", "checked_at"])
        hist_df.rename(
            columns={"name": "Товар", "price": "Цена (₽)", "checked_at": "Когда проверено"},
            inplace=True,
        )
        hist_df.to_excel(xw, sheet_name="История", index=False)

    # === Стилизация ===
    wb = openpyxl.load_workbook(out_path)
    ws = wb["Текущие цены"]

    # зафиксировать заголовок
    ws.freeze_panes = "A2"

    # автоширина колонок по содержимому
    def autosize(ws, padding=2, min_w=12, max_w=60):
        dims = {}
        for row in ws.iter_rows(values_only=True):
            for i, cell in enumerate(row, 1):
                if cell is not None:
                    dims[i] = max(dims.get(i, 0), len(str(cell)))
        for col_idx, width in dims.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + padding, min_w), max_w)

    autosize(ws)

    # заголовки — жирные, по центру, перенос
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # сделать колонку «Товар» широкой по максимальной длине наименования
    headers = {cell.value: cell.column for cell in ws[1]}
    col_product = headers.get("Товар")
    if col_product:
        max_name_len = 0
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=col_product).value
            if v:
                max_name_len = max(max_name_len, len(str(v)))
        if max_name_len:
            ws.column_dimensions[get_column_letter(col_product)].width = min(max(max_name_len + 4, 24), 70)

    # применить числовые форматы и подсветку
    col_change = headers.get("Изменение (₽)")
    col_price = headers.get("Цена (₽)")
    col_prev = headers.get("Предыдущая цена (₽)")
    col_change_pct = headers.get("Изменение (%)")

    if col_price:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_price).number_format = "#,##0"
    if col_prev:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_prev).number_format = "#,##0"
    if col_change:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_change).number_format = "#,##0"
    if col_change_pct:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_change_pct).number_format = "0.00"

    # подсветка изменений
    if col_change:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_change)
            prev_cell = ws.cell(row=row, column=col_prev) if col_prev else None
            price_cell = ws.cell(row=row, column=col_price) if col_price else None
            try:
                delta = float(cell.value)
            except (TypeError, ValueError):
                delta = 0
            if delta < 0:
                # зелёный — цена снизилась
                fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
                for c in (cell, prev_cell, price_cell):
                    if c is not None:
                        c.font = Font(bold=True)
                        c.fill = fill
            elif delta > 0:
                # розовый — цена выросла
                fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
                cell.fill = fill

    # Стили для «История»
    if "История" in wb.sheetnames:
        wh = wb["История"]
        wh.freeze_panes = "A2"
        autosize(wh)
        for c in wh[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wh.row_dimensions[1].height = 28
        # форматы чисел
        hdrs_h = {cell.value: cell.column for cell in wh[1]}
        c_price_h = hdrs_h.get("Цена (₽)")
        if c_price_h:
            for row in range(2, wh.max_row + 1):
                wh.cell(row=row, column=c_price_h).number_format = "#,##0"

    # Лист «Легенда» с объяснениями
    sheet_name = "Легенда"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    lg = wb.create_sheet(sheet_name)
    lg["A1"] = "Обозначения и подсветка"
    lg["A1"].font = Font(bold=True)
    lg["A3"] = "• «Изменение (₽)» — разница между текущей ценой и предыдущей."
    lg["A4"] = "• «Изменение (%)» — отношение изменения к предыдущей цене × 100."
    lg["A5"] = "• Зелёная подсветка — цена снизилась."
    lg["A6"] = "• Розовая подсветка — цена выросла."
    lg["A7"] = "• Если предыдущей цены нет (первый замер), изменение не вычисляется."

    autosize(lg, min_w=30)
    wb.save(out_path)
    return out_path
